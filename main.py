import csv
import re
from difflib import SequenceMatcher
from pathlib import Path

try:
	from openpyxl import load_workbook
except ModuleNotFoundError as error:
	raise SystemExit(
		"Missing dependency: openpyxl. Install it with: python -m pip install -r requirements.txt"
	) from error


CSV_FILE = Path(__file__).with_name("MaterialList.csv")
QUERY_FILE = Path(__file__).with_name("MaterialQuery.xlsx")
RESULT_FILE = Path(__file__).with_name("result.csv")
QUERY_COLUMN_NAMES = {"materialdescription", "material description"}
MATCH_LIMIT = 5


def normalize_text(value):
	return " ".join(re.findall(r"[a-z0-9]+", value.lower()))


def score_description(query, description):
	normalized_query = normalize_text(query)
	normalized_description = normalize_text(description)

	if not normalized_query or not normalized_description:
		return 0.0

	if normalized_query in normalized_description:
		return 1.0

	query_words = set(normalized_query.split())
	description_words = set(normalized_description.split())
	word_overlap = len(query_words & description_words) / len(query_words)
	text_similarity = SequenceMatcher(None, normalized_query, normalized_description).ratio()

	return (word_overlap * 0.7) + (text_similarity * 0.3)


def load_materials(csv_file):
	with csv_file.open(newline="", encoding="utf-8-sig") as file:
		reader = csv.DictReader(file)
		required_columns = {"MaterialDescription", "Material"}
		missing_columns = required_columns - set(reader.fieldnames or [])

		if missing_columns:
			missing = ", ".join(sorted(missing_columns))
			raise ValueError(f"Missing required column(s): {missing}")

		return [
			{
				"description": row["MaterialDescription"].strip(),
				"material": row["Material"].strip(),
			}
			for row in reader
			if row.get("MaterialDescription") and row.get("Material")
		]


def load_queries(query_file):
	workbook = load_workbook(query_file, read_only=True, data_only=True)
	worksheet = workbook.active
	headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
	query_column = None

	for index, header in enumerate(headers, start=1):
		if header and str(header).strip().lower() in QUERY_COLUMN_NAMES:
			query_column = index
			break

	if query_column is None:
		raise ValueError(
			"MaterialQuery.xlsx must contain a 'Material Description' column."
		)

	queries = []
	for row_number, row in enumerate(
		worksheet.iter_rows(min_row=2, values_only=True), start=2
	):
		query = row[query_column - 1]
		if query is None or str(query).strip() == "":
			continue

		queries.append({"row": row_number, "description": str(query).strip()})

	return queries


def find_top_materials(query, materials, limit=5):
	ranked_matches = []
	seen = set()

	for item in materials:
		key = (item["material"], item["description"])
		if key in seen:
			continue

		seen.add(key)
		score = score_description(query, item["description"])
		ranked_matches.append((item, score))

	ranked_matches.sort(key=lambda match: match[1], reverse=True)
	return ranked_matches[:limit]


def save_results(results, result_file):
	fieldnames = [
		"QueryRow",
		"QueryMaterialDescription",
		"MatchRank",
		"Material",
		"MaterialDescription",
		"MatchScore",
	]

	with result_file.open("w", newline="", encoding="utf-8-sig") as file:
		writer = csv.DictWriter(file, fieldnames=fieldnames)
		writer.writeheader()

		for result in results:
			writer.writerow(
				{
					"QueryRow": result["query_row"],
					"QueryMaterialDescription": result["query_description"],
					"MatchRank": result["rank"],
					"Material": result["material"],
					"MaterialDescription": result["description"],
					"MatchScore": round(result["score"], 4),
				}
			)


def main():
	materials = load_materials(CSV_FILE)
	queries = load_queries(QUERY_FILE)
	results = []

	for query in queries:
		matches = find_top_materials(query["description"], materials, MATCH_LIMIT)
		for rank, (match, score) in enumerate(matches, start=1):
			results.append(
				{
					"query_row": query["row"],
					"query_description": query["description"],
					"rank": rank,
					"material": match["material"],
					"description": match["description"],
					"score": score,
				}
			)

	save_results(results, RESULT_FILE)
	print(f"Saved {len(results)} matching rows to {RESULT_FILE.name}")


if __name__ == "__main__":
	main()
